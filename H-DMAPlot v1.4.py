# H-DMAPlot GUI v1.4
# Análise Termo-Mecânica · Seleção Dinâmica de Colunas
# Autor: Carlos Henrique Amaro da Silva

import os
import re
import io
import math
import tempfile
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure

import tkinter as tk
from tkinter import filedialog, messagebox, ttk, colorchooser

from scipy.signal import savgol_filter, find_peaks
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# RESOURCE PATH (PyInstaller) 
import sys

def resource_path(relative_path):
    """Resolve o caminho de recursos — funciona tanto em desenvolvimento
    quanto empacotado com PyInstaller (--onefile ou --onedir)."""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


# PALETA 

BG        = "#0f1117"
SURFACE   = "#1a1d27"
CARD      = "#22263a"
ACCENT    = "#4f8ef7"
ACCENT2   = "#7c3aed"
SUCCESS   = "#22c55e"
WARNING   = "#f59e0b"
ERROR     = "#ef4444"
TEXT      = "#e2e8f0"
TEXT_DIM  = "#64748b"
BORDER    = "#2d3148"

FONT_TITLE  = ("Segoe UI", 16, "bold")
FONT_HEAD   = ("Segoe UI", 11, "bold")
FONT_BODY   = ("Segoe UI", 10)
FONT_SMALL  = ("Segoe UI", 9)
FONT_MONO   = ("Consolas", 9)

CORES_TAB10 = [
    "#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd",
    "#8c564b", "#e377c2", "#7f7f7f", "#bcbd22", "#17becf",
]

ESTILOS_LINHA = {
    "Sólida":      "-",
    "Tracejada":   "--",
    "Pontilhada":  ":",
    "Traço-ponto": "-.",
}

# MAPEAMENTO DE COLUNAS DMA
# Chave: nome da coluna (case-insensitive, sem unidade)
# Valor: descrição amigável exibida antes do rótulo técnico
DESCRICOES_COLUNAS = {
    # Identificação / tempo
    "index":   "Índice do ponto",
    "ts":      "Temperatura da amostra",
    "t":       "Tempo",
    "f":       "Frequência de excitação",        # f minúsculo → frequência [Hz]
    "tr":      "Temperatura de referência",

    # Forças e deslocamentos
    "f0":      "Força de excitação (referência)",
    "x0":      "Deslocamento de referência",
    "F":       "Força instantânea aplicada",     # F maiúsculo → força [N]
    "x":       "Deslocamento dinâmico",
    "phase":   "Ângulo de fase (δ)",

    # Módulos elásticos (armazenamento / perda / complexo)
    "e'":      "Módulo de Armazenamento (elástico)",
    "e''":     "Módulo de Perda (viscoso)",
    "e\"":     "Módulo de Perda (viscoso)",
    "e*":      "Módulo Complexo (|E*|)",

    # Amortecimento
    "tan_delta": "Fator de Amortecimento / Fator de Perda",

    # Compliance (inverso do módulo)
    "d'":      "Compliance de Armazenamento",
    "d''":     "Compliance de Perda",
    "d\"":     "Compliance de Perda",
    "d*":      "Compliance Complexo",

    # Viscosidade dinâmica
    "eta'":    "Viscosidade Dinâmica (componente real)",
    "eta''":   "Viscosidade Dinâmica (componente imaginária)",
    "eta\"":   "Viscosidade Dinâmica (componente imaginária)",
    "eta*":    "Viscosidade Complexa (|η*|)",
}

def _descricao_coluna(col_label: str) -> str:
    """
    Dado um rótulo como "E' [MPa]" ou "tan_delta []",
    retorna string no formato "Módulo de Armazenamento — E' [MPa]".
    Se não houver descrição conhecida, retorna o rótulo original.
    """
    # extrai só o nome (antes do espaço + unidade entre colchetes)
    # NÃO converte para lower() para preservar distinção entre 'f' (frequência) e 'F' (força)
    nome = col_label.split(" [")[0].strip()
    desc = DESCRICOES_COLUNAS.get(nome) or DESCRICOES_COLUNAS.get(nome.lower())
    if desc:
        return f"{desc}  —  {col_label}"
    return col_label

# PARSING  

def _parse_txt(caminho: str) -> pd.DataFrame | None:
    """
    Lê arquivo TXT com cabeçalho de 2 linhas (nomes + unidades) e dados numéricos.
    Suporta separador decimal vírgula ou ponto; ignora linhas inválidas e marcadores '&'.
    Retorna DataFrame com colunas no formato 'Nome [unidade]'.
    """
    # Tenta UTF-8 primeiro; se falhar, cai para latin-1 (ISO-8859-1)
    # que cobre a maioria dos arquivos de equipamentos europeus (°C, µm, etc.)
    for enc in ("utf-8-sig", "latin-1"):
        try:
            with open(caminho, "r", encoding=enc) as f:
                linhas = f.readlines()
            break
        except UnicodeDecodeError:
            continue
    else:
        with open(caminho, "r", encoding="utf-8", errors="replace") as f:
            linhas = f.readlines()

    # Localiza as duas primeiras linhas não-vazias que servem de cabeçalho
    cab_idx = []
    for i, ln in enumerate(linhas):
        stripped = ln.strip()
        if stripped and not stripped.startswith("&") and not stripped.startswith("#"):
            cab_idx.append(i)
        if len(cab_idx) == 2:
            break

    if len(cab_idx) < 2:
        return None

    def _split(ln):
        return ln.strip().split()

    nomes   = _split(linhas[cab_idx[0]])
    unidades = _split(linhas[cab_idx[1]])
    # Garante mesmo comprimento
    while len(unidades) < len(nomes):
        unidades.append("[]")

    # Determina quais índices manter (exclui colunas de índice sequencial sem valor físico)
    _IGNORAR_NOMES = {"index", "idx"}
    _IGNORAR_UNIDADES = {"[#]"}
    indices_manter = [
        i for i, (n, u) in enumerate(zip(nomes, unidades))
        if n.lower() not in _IGNORAR_NOMES and u.lower() not in _IGNORAR_UNIDADES
    ]
    col_labels = [f"{nomes[i]} {unidades[i]}" for i in indices_manter]

    # Le linhas de dados e deixa a conversao numerica para o pandas.
    data_lines = []
    for ln in linhas[cab_idx[1]+1:]:
        stripped = ln.strip()
        if not stripped or stripped.startswith("&") or stripped.startswith("#"):
            continue
        partes = stripped.split()
        if len(partes) < len(nomes):
            continue
        data_lines.append(" ".join(partes[:len(nomes)]).replace(",", "."))

    if not data_lines:
        return None

    try:
        df = pd.read_csv(
            io.StringIO("\n".join(data_lines)),
            sep=r"\s+",
            header=None,
            usecols=indices_manter,
            engine="python",
            on_bad_lines="skip",
        )
    except Exception:
        return None

    df = df.apply(pd.to_numeric, errors="coerce").dropna(how="any")
    if df.empty:
        return None
    df.columns = col_labels
    return df.reset_index(drop=True)


# MODELO 

class Amostra:
    _contador = 0

    def __init__(self, nome: str, caminho: str, df: pd.DataFrame):
        self.nome      = nome
        self.caminho   = caminho
        self.df        = df
        self.linestyle = "-"
        self.linewidth = 1.8
        idx = Amostra._contador % len(CORES_TAB10)
        self.color     = CORES_TAB10[idx]
        Amostra._contador += 1
        # tratamento: offset por coluna {col_label: offset}
        self.offsets: dict[str, float] = {}
        # janela de corte por coluna {col_label: (i_ini, i_fim)}
        self.cortes: dict[str, tuple[int, int]] = {}
        # suavização por coluna {col_label: {"metodo": str, "janela": int, "poly": int}}
        self.suavizacoes: dict[str, dict] = {}
        # Cache de series tratadas para evitar recalcular suavizacao em redesenhos repetidos.
        self._serie_cache: dict[tuple, np.ndarray] = {}

    def get_serie(self, col: str) -> np.ndarray:
        """Retorna série com offset, corte e suavização aplicados."""
        corte = self.cortes.get(col)
        suav = self.suavizacoes.get(col)
        suav_key = None
        if suav:
            suav_key = (
                suav.get("metodo", ""),
                int(suav.get("janela", 5)),
                int(suav.get("poly", 3)),
            )
        cache_key = (col, corte, suav_key, float(self.offsets.get(col, 0.0)))
        cached = self._serie_cache.get(cache_key)
        if cached is not None:
            return cached

        vals = self.df[col].to_numpy(dtype=float)
        # corte
        if corte is not None:
            i0, i1 = corte
            vals = vals[i0:i1+1]
        # suavização
        if suav:
            cfg = suav
            metodo = cfg.get("metodo", "")
            janela = int(cfg.get("janela", 5))
            janela = max(3, janela if janela % 2 == 1 else janela + 1)  # deve ser ímpar
            if metodo == "savgol" and len(vals) > janela:
                poly = min(int(cfg.get("poly", 3)), janela - 1)
                try:
                    vals = savgol_filter(vals, janela, poly)
                except Exception:
                    pass
            elif metodo == "media_movel" and len(vals) > janela:
                kernel = np.ones(janela) / janela
                vals = np.convolve(vals, kernel, mode="same")
        # offset
        offset = self.offsets.get(col, 0.0)
        vals = vals + offset
        self._serie_cache[cache_key] = vals
        return vals

    def invalidar_cache(self):
        self._serie_cache.clear()

    @property
    def colunas(self) -> list[str]:
        return list(self.df.columns)


# GUI HELPERS 

def btn(parent, text, cmd, color=ACCENT, fg="white", **kw):
    b = tk.Button(parent, text=text, command=cmd, bg=color, fg=fg,
                  font=FONT_SMALL, relief="flat", cursor="hand2",
                  activebackground=_lighten(color), activeforeground=fg,
                  padx=10, pady=5, **kw)
    b.bind("<Enter>", lambda e: b.config(bg=_lighten(color)))
    b.bind("<Leave>", lambda e: b.config(bg=color))
    return b

def _lighten(hex_color, amount=30):
    hex_color = hex_color.lstrip("#")
    r, g, b_ = (int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    return f"#{min(255,r+amount):02x}{min(255,g+amount):02x}{min(255,b_+amount):02x}"

def lbl(parent, text, font=FONT_BODY, fg=TEXT, **kw):
    return tk.Label(parent, text=text, font=font, fg=fg, bg=parent["bg"], **kw)

def card(parent, **kw):
    return tk.Frame(parent, bg=CARD, bd=0, highlightthickness=1,
                    highlightbackground=BORDER, **kw)

def sep(parent):
    return tk.Frame(parent, bg=BORDER, height=1)

def scrolled_frame(parent, bg=SURFACE):
    """Retorna (outer_frame, inner_frame) com scrollbar vertical."""
    outer = tk.Frame(parent, bg=bg)
    canvas_scroll = tk.Canvas(outer, bg=bg, highlightthickness=0)
    sb = tk.Scrollbar(outer, orient="vertical", command=canvas_scroll.yview)
    canvas_scroll.configure(yscrollcommand=sb.set)
    sb.pack(side="right", fill="y")
    canvas_scroll.pack(side="left", fill="both", expand=True)
    inner = tk.Frame(canvas_scroll, bg=bg)
    window_id = canvas_scroll.create_window((0, 0), window=inner, anchor="nw")

    def _on_configure(event):
        canvas_scroll.configure(scrollregion=canvas_scroll.bbox("all"))
        canvas_scroll.itemconfig(window_id, width=canvas_scroll.winfo_width())

    inner.bind("<Configure>", _on_configure)
    canvas_scroll.bind("<Configure>", _on_configure)

    def _mousewheel(event):
        canvas_scroll.yview_scroll(int(-1*(event.delta/120)), "units")

    canvas_scroll.bind_all("<MouseWheel>", _mousewheel)
    return outer, inner


# POPUP EDITOR

def popup_edit(root, titulo, valor_atual, callback):
    popup = tk.Toplevel(root)
    popup.title(titulo)
    popup.configure(bg="white")
    popup.resizable(False, False)
    popup.grab_set()
    popup.attributes("-topmost", True)
    pw, ph = 380, 115
    sw, sh = popup.winfo_screenwidth(), popup.winfo_screenheight()
    popup.geometry(f"{pw}x{ph}+{(sw-pw)//2}+{(sh-ph)//2}")

    tk.Label(popup, text=titulo, font=("Segoe UI", 9, "bold"),
             fg="#555", bg="white").pack(anchor="w", padx=14, pady=(10, 2))

    var = tk.StringVar(value=valor_atual)
    entry = tk.Entry(popup, textvariable=var, font=("Segoe UI", 11),
                     bg="white", fg="#111", insertbackground="#111",
                     relief="flat", highlightthickness=1,
                     highlightcolor=ACCENT, highlightbackground="#ccc")
    entry.pack(fill="x", padx=14, pady=4, ipady=5)
    entry.select_range(0, "end")
    entry.focus_set()

    def _ok(event=None):
        novo = var.get().strip()
        popup.destroy()
        if novo:
            callback(novo)

    def _cancel(event=None):
        popup.destroy()

    bf = tk.Frame(popup, bg="white")
    bf.pack(anchor="e", padx=14, pady=(2, 10))
    tk.Button(bf, text="Cancelar", command=_cancel,
              bg="#f1f5f9", fg="#555", font=("Segoe UI", 9),
              relief="flat", padx=10, pady=4, cursor="hand2").pack(side="left", padx=4)
    tk.Button(bf, text="  OK  ", command=_ok,
              bg=ACCENT, fg="white", font=("Segoe UI", 9, "bold"),
              relief="flat", padx=10, pady=4, cursor="hand2").pack(side="left")

    entry.bind("<Return>", _ok)
    entry.bind("<Escape>", _cancel)
    popup.wait_window()


# CANVAS DO GRÁFICO 

class PlotCanvas:
    def __init__(self, parent):
        self.parent = parent
        self.fig = Figure(figsize=(10, 5), dpi=96)
        self.fig.patch.set_facecolor("white")
        self.canvas = FigureCanvasTkAgg(self.fig, master=parent)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)
        toolbar_frame = tk.Frame(parent, bg="white")
        toolbar_frame.pack(fill="x")
        self.toolbar = NavigationToolbar2Tk(self.canvas, toolbar_frame)
        self.toolbar.config(bg="white")
        self.toolbar.update()

        # metadados editáveis
        self.titulo   = ""
        self.xlabel   = ""
        self.ylabels: list[str] = []  # um por eixo Y selecionado
        self._cid     = None
        self._axes: list = []
        self._line_refs: list[list] = []  # [eixo_y_idx][amostra_idx]
        self._crosshair_series: list[dict] = []
        self._crosshair_lines: list = []
        self._crosshair_tooltip = None
        self._cid_cross_motion = None
        self._cid_cross_click = None
        self._fixed_crosshair_marks: list[dict] = []
        self._next_fixed_crosshair_id: int = 1

        # Configurações de detecção de picos 
        self.picos_ativos: bool = False          # ligar/desligar picos
        self.picos_prominence: float = 0.05     # proeminência mínima (fração do range Y)
        self.picos_distancia: int = 10           # distância mínima entre picos (pontos)
        self.picos_max: int = 5                  # máximo de picos por curva

        # Configurações de fontes 
        self.fonte_titulo:  int = 11
        self.fonte_eixo:    int = 10
        self.fonte_ticks:   int = 9
        self.fonte_legenda: int = 9
        self.fonte_picos:   int = 8

        # Configurações de grid 
        self.grid_x:         bool  = True   # linhas verticais (sobre eixo X)
        self.grid_y:         bool  = True   # linhas horizontais (sobre eixo Y)
        self.grid_intervalo: float = 0.0    # 0 = automático; >0 = passo fixo em unid. X
        self.xlim_min: float | None = None
        self.xlim_max: float | None = None
        self.sobrepor_curvas: bool = False   # todas as curvas Y no mesmo eixo

        # Estado para arrastar anotações 
        # lista de dicts: {annot, ax, x_data, y_data}
        self._annots: list[dict] = []
        self._drag_annot = None    # anotação sendo arrastada
        self._drag_offset = (0, 0) # offset (dx, dy) em coordenadas de dados
        self._cid_press   = None
        self._cid_release = None
        self._cid_motion  = None

        self._mostrar_placeholder()

    # Helpers de arrasto 

    def _conectar_drag(self):
        """Conecta eventos de mouse para arrastar anotações de pico."""
        self._desconectar_drag()
        self._cid_press   = self.canvas.mpl_connect("button_press_event",   self._on_drag_press)
        self._cid_release = self.canvas.mpl_connect("button_release_event", self._on_drag_release)
        self._cid_motion  = self.canvas.mpl_connect("motion_notify_event",  self._on_drag_motion)

    def _desconectar_drag(self):
        for cid_attr in ("_cid_press", "_cid_release", "_cid_motion"):
            cid = getattr(self, cid_attr, None)
            if cid is not None:
                try: self.canvas.mpl_disconnect(cid)
                except Exception: pass
            setattr(self, cid_attr, None)

    def _toolbar_ativa(self) -> bool:
        return bool(getattr(self.toolbar, "mode", ""))

    @staticmethod
    def _pixel_para_dados(ax, x_pixel, y_pixel):
        return ax.transData.inverted().transform((x_pixel, y_pixel))

    def _evento_sobre_anotacao(self, event) -> bool:
        try:
            renderer = self.canvas.get_renderer()
        except Exception:
            return False
        for info in self._annots:
            ann = info["annot"]
            try:
                bbox = ann.get_window_extent(renderer)
            except Exception:
                continue
            pad = 6
            if (bbox.x0 - pad <= event.x <= bbox.x1 + pad and
                    bbox.y0 - pad <= event.y <= bbox.y1 + pad):
                return True
        return False

    def _on_drag_press(self, event):
        if event.button != 1:
            return
        # Procura anotação próxima ao clique (em pixels)
        try:
            renderer = self.canvas.get_renderer()
        except Exception:
            return
        for info in self._annots:
            ann = info["annot"]
            bbox = ann.get_window_extent(renderer)
            # Expande levemente a bbox para facilitar o clique
            pad = 6
            if (bbox.x0 - pad <= event.x <= bbox.x1 + pad and
                    bbox.y0 - pad <= event.y <= bbox.y1 + pad):
                self._drag_annot = info
                # offset em coordenadas de dados entre o ponto do clique e a posição atual da anotação
                ax = info["ax"]
                ann_xy = ann.get_position()       # posição em dados
                click_xy = self._pixel_para_dados(ax, event.x, event.y)
                self._drag_offset = (ann_xy[0] - click_xy[0], ann_xy[1] - click_xy[1])
                return

    def _on_drag_release(self, event):
        self._drag_annot = None

    def _on_drag_motion(self, event):
        if self._drag_annot is None:
            return
        info = self._drag_annot
        ax = info["ax"]
        event_x, event_y = self._pixel_para_dados(ax, event.x, event.y)
        ann  = info["annot"]
        novo_x = event_x + self._drag_offset[0]
        novo_y = event_y + self._drag_offset[1]
        ann.set_position((novo_x, novo_y))
        # Atualiza a seta para apontar ao ponto original do pico
        ann.xy = (info["x_data"], info["y_data"])
        mark_id = info.get("mark_id")
        serie_key = info.get("serie_key")
        if mark_id is not None and serie_key is not None:
            mark = next((m for m in self._fixed_crosshair_marks if m["id"] == mark_id), None)
            if mark is not None:
                mark.setdefault("offsets", {})[serie_key] = (
                    novo_x - info["x_data"],
                    novo_y - info["y_data"],
                )
        self.canvas.draw_idle()

    def _desconectar_crosshair(self):
        for cid_attr in ("_cid_cross_motion", "_cid_cross_click"):
            cid = getattr(self, cid_attr, None)
            if cid is not None:
                try: self.canvas.mpl_disconnect(cid)
                except Exception: pass
            setattr(self, cid_attr, None)

    def _limpar_crosshair(self):
        for line in self._crosshair_lines:
            try: line.remove()
            except Exception: pass
        self._crosshair_lines = []
        if self._crosshair_tooltip is not None:
            try: self._crosshair_tooltip.remove()
            except Exception: pass
            self._crosshair_tooltip = None

    @staticmethod
    def _interp_em_x(x_vals, y_vals, x_alvo):
        x = np.asarray(x_vals, dtype=float)
        y = np.asarray(y_vals, dtype=float)
        mask = np.isfinite(x) & np.isfinite(y)
        x = x[mask]
        y = y[mask]
        if len(x) == 0:
            return None

        ordem = np.argsort(x)
        x = x[ordem]
        y = y[ordem]
        x_unico, idx_unico = np.unique(x, return_index=True)
        y_unico = y[idx_unico]
        if len(x_unico) == 0 or x_alvo < x_unico[0] or x_alvo > x_unico[-1]:
            return None
        return float(np.interp(x_alvo, x_unico, y_unico))

    def _valores_crosshair(self, x_alvo):
        valores = []
        for serie in self._crosshair_series:
            y_alvo = self._interp_em_x(serie["x"], serie["y"], x_alvo)
            if y_alvo is None:
                continue
            valores.append({
                "ax": serie["ax"],
                "serie_key": serie["serie_key"],
                "label": serie["label"],
                "color": serie["color"],
                "x": float(x_alvo),
                "y": y_alvo,
            })
        return valores

    def _desenhar_crosshair_fixo(self, mark):
        x_alvo = float(mark["x"])
        valores = self._valores_crosshair(x_alvo)
        if not valores:
            return 0

        offsets = mark.setdefault("offsets", {})
        total = 0
        for idx, item in enumerate(valores):
            ax_ref = item["ax"]
            cor = item["color"]
            y_min, y_max = ax_ref.get_ylim()
            y_range = (y_max - y_min) or 1.0
            x_min, x_max = ax_ref.get_xlim()
            x_range = (x_max - x_min) or 1.0

            serie_key = item["serie_key"]
            if serie_key in offsets:
                dx, dy = offsets[serie_key]
            else:
                dx = x_range * (0.015 + 0.012 * (idx % 3))
                dy = y_range * (0.08 + 0.045 * (idx % 4))
                offsets[serie_key] = (dx, dy)

            texto = f"{x_alvo:.1f}\n({item['y']:.3g})"
            ann = ax_ref.annotate(
                texto,
                xy=(x_alvo, item["y"]),
                xytext=(x_alvo + dx, item["y"] + dy),
                fontsize=max(7, self.fonte_picos),
                color=cor,
                ha="center",
                va="bottom",
                bbox=dict(
                    boxstyle="round,pad=0.25",
                    facecolor="white",
                    edgecolor=cor,
                    alpha=0.9,
                    linewidth=0.8,
                ),
                arrowprops=dict(
                    arrowstyle="-|>",
                    color=cor,
                    lw=0.9,
                    mutation_scale=7,
                ),
                annotation_clip=False,
                zorder=25,
                picker=True,
            )
            ax_ref.plot(
                x_alvo, item["y"], marker="o", color=cor,
                markersize=4.5, zorder=24, linestyle="none"
            )
            self._annots.append({
                "annot": ann,
                "ax": ax_ref,
                "x_data": x_alvo,
                "y_data": item["y"],
                "amostra": "crosshair",
                "mark_id": mark["id"],
                "serie_key": serie_key,
            })
            total += 1
        return total

    def _texto_crosshair(self, x_alvo, valores):
        linhas = [f"{self.xlabel or 'X'} = {x_alvo:.2f}"]
        for item in valores:
            linhas.append(f"{item['label']}: {item['y']:.4g}")
        return "\n".join(linhas)

    def _garantir_crosshair_artistas(self):
        if self._crosshair_lines or not self._axes:
            return
        for ax in self._axes:
            line = ax.axvline(
                0, color="#334155", linestyle="--", linewidth=0.9,
                alpha=0.65, visible=False, zorder=20
            )
            self._crosshair_lines.append(line)

    def _on_crosshair_motion(self, event):
        if self._toolbar_ativa() or self._drag_annot is not None:
            return
        if event.inaxes not in self._axes or event.xdata is None:
            self._limpar_crosshair()
            self.canvas.draw_idle()
            return

        x_alvo = float(event.xdata)
        valores = self._valores_crosshair(x_alvo)
        if not valores:
            self._limpar_crosshair()
            self.canvas.draw_idle()
            return
        item_ref = next((v for v in valores if v["ax"] is event.inaxes), valores[0])

        self._garantir_crosshair_artistas()
        for line in self._crosshair_lines:
            line.set_xdata([x_alvo, x_alvo])
            line.set_visible(True)

        if (self._crosshair_tooltip is not None and
                getattr(self._crosshair_tooltip, "axes", None) is not event.inaxes):
            try: self._crosshair_tooltip.remove()
            except Exception: pass
            self._crosshair_tooltip = None

        if self._crosshair_tooltip is None:
            self._crosshair_tooltip = event.inaxes.annotate(
                "",
                xy=(x_alvo, item_ref["y"]),
                xytext=(12, 12),
                textcoords="offset points",
                fontsize=max(7, self.fonte_picos),
                color="#111827",
                ha="left",
                va="bottom",
                bbox=dict(
                    boxstyle="round,pad=0.35",
                    facecolor="white",
                    edgecolor="#334155",
                    alpha=0.94,
                    linewidth=0.8,
                ),
                annotation_clip=False,
                zorder=30,
            )
        self._crosshair_tooltip.xy = (x_alvo, item_ref["y"])
        self._crosshair_tooltip.set_text(self._texto_crosshair(x_alvo, valores))
        self._crosshair_tooltip.set_visible(True)
        self.canvas.draw_idle()

    def _fixar_crosshair(self, event):
        if self._toolbar_ativa() or self._drag_annot is not None:
            return
        if event.button != 1 or event.dblclick or event.inaxes not in self._axes or event.xdata is None:
            return
        if self._evento_sobre_anotacao(event):
            return

        x_alvo = float(event.xdata)
        valores = self._valores_crosshair(x_alvo)
        if not valores:
            return

        mark = {
            "id": self._next_fixed_crosshair_id,
            "x": x_alvo,
            "offsets": {},
        }
        self._next_fixed_crosshair_id += 1
        self._fixed_crosshair_marks.append(mark)
        self._desenhar_crosshair_fixo(mark)
        self._conectar_drag()
        self.canvas.draw_idle()

    def _conectar_crosshair(self):
        self._desconectar_crosshair()
        if not self._crosshair_series:
            return
        self._cid_cross_motion = self.canvas.mpl_connect(
            "motion_notify_event", self._on_crosshair_motion
        )
        self._cid_cross_click = self.canvas.mpl_connect(
            "button_press_event", self._fixar_crosshair
        )

    def limpar_marcacoes_fixas(self):
        self._fixed_crosshair_marks.clear()
        self._limpar_crosshair()

    @staticmethod
    def _estilizar(ax):
        ax.set_facecolor("white")
        ax.tick_params(colors="#111", labelsize=9)
        ax.xaxis.label.set_color("#111")
        ax.yaxis.label.set_color("#111")
        ax.title.set_color("#111")
        for sp in ax.spines.values():
            sp.set_visible(True)
            sp.set_edgecolor("#aaa")
            sp.set_linewidth(0.8)

    def _desenhar_picos(self, ax, x_vals, y_vals, cor, nome_amostra):
        """
        Detecta picos em (x_vals, y_vals) e plota anotações arrastáveis.
        Retorna lista de dicts para _annots.
        """
        if len(y_vals) < 3:
            return []

        y_range = float(np.ptp(y_vals))
        if y_range == 0:
            return []

        prominence_abs = self.picos_prominence * y_range
        distancia = max(1, int(self.picos_distancia))

        try:
            idx_picos, props = find_peaks(
                y_vals,
                prominence=prominence_abs,
                distance=distancia,
            )
        except Exception:
            return []

        if len(idx_picos) == 0:
            return []

        # Ordena por proeminência e mantém só os top N
        prominencias = props.get("prominences", np.ones(len(idx_picos)))
        ordem = np.argsort(prominencias)[::-1]
        idx_picos = idx_picos[ordem[:self.picos_max]]

        novos_annots = []
        for ip in idx_picos:
            xp = float(x_vals[ip])
            yp = float(y_vals[ip])

            # Offset padrão da anotação: um pouco acima e à direita do pico
            x_off = 0.0
            y_off = y_range * 0.08

            ann = ax.annotate(
                f"{xp:.1f}\n({yp:.3g})",
                xy=(xp, yp),
                xytext=(xp + x_off, yp + y_off),
                fontsize=self.fonte_picos,
                color=cor,
                ha="center",
                va="bottom",
                bbox=dict(
                    boxstyle="round,pad=0.25",
                    facecolor="white",
                    edgecolor=cor,
                    alpha=0.88,
                    linewidth=0.8,
                ),
                arrowprops=dict(
                    arrowstyle="-|>",
                    color=cor,
                    lw=0.9,
                    mutation_scale=7,
                ),
                annotation_clip=False,
                zorder=10,
                picker=True,
            )
            ax.plot(xp, yp, marker="v", color=cor, markersize=6,
                    zorder=9, linestyle="none")

            novos_annots.append({
                "annot": ann,
                "ax": ax,
                "x_data": xp,
                "y_data": yp,
                "amostra": nome_amostra,
            })

        return novos_annots

    def _mostrar_placeholder(self):
        self.fig.clear()
        ax = self.fig.add_subplot(111)
        self._estilizar(ax)
        ax.text(0.5, 0.5,
                "Carregue arquivos .txt e configure\nos eixos no painel esquerdo",
                ha="center", va="center", transform=ax.transAxes,
                fontsize=12, color="#aaa", fontfamily="Segoe UI")
        ax.set_xticks([]); ax.set_yticks([])
        for sp in ax.spines.values(): sp.set_visible(False)
        self.canvas.draw_idle()

    def _criar_eixos_sobrepostos(self, n_y: int):
        """Cria um eixo Y independente para cada curva selecionada."""
        host = self.fig.add_subplot(111)
        axes = [host]
        desloc = 54

        for j in range(1, n_y):
            ax = host.twinx()
            if j % 2 == 0:
                ax.spines["left"].set_position(("outward", desloc * (j // 2)))
                ax.spines["left"].set_visible(True)
                ax.spines["right"].set_visible(False)
                ax.yaxis.set_label_position("left")
                ax.yaxis.tick_left()
            else:
                ax.spines["right"].set_position(("outward", desloc * ((j + 1) // 2 - 1)))
                ax.spines["right"].set_visible(True)
                ax.yaxis.set_label_position("right")
                ax.yaxis.tick_right()
            axes.append(ax)

        return axes

    def redesenhar(self, amostras: list[Amostra], col_x: str,
                   cols_y: list[str], mostrar_grid: bool = True,
                   root_for_popup=None, sobrepor_curvas: bool = False):
        if not amostras or not col_x or not cols_y:
            self._desconectar_crosshair()
            self._mostrar_placeholder()
            return

        self.fig.clear()
        self._annots = []   # limpa anotações anteriores
        self._crosshair_series = []
        self._crosshair_lines = []
        self._crosshair_tooltip = None
        n_y = len(cols_y)
        self.sobrepor_curvas = sobrepor_curvas

        # Um subplot por coluna Y, ou todas as curvas no mesmo eixo quando sobreposto.
        if sobrepor_curvas:
            axes = self._criar_eixos_sobrepostos(n_y)
        else:
            axes = self.fig.subplots(1, n_y) if n_y > 1 else [self.fig.add_subplot(111)]
            if n_y == 1:
                axes = [axes[0]] if not isinstance(axes, list) else axes
            else:
                axes = list(axes)

        self._axes = axes
        self._line_refs = [[] for _ in range(n_y)]

        for ax in axes:
            self._estilizar(ax)

        for i, amostra in enumerate(amostras):
            if col_x not in amostra.df.columns:
                continue
            for j, col_y in enumerate(cols_y):
                if col_y not in amostra.df.columns:
                    continue
                # X e Y passam ambos por get_serie() para respeitar corte/offset/suavização
                x_vals = amostra.get_serie(col_x)
                y_vals = amostra.get_serie(col_y)
                # Se apenas Y tem corte definido (e X não), alinhar fatia de X ao corte de Y
                if col_y in amostra.cortes and col_x not in amostra.cortes:
                    i0, i1 = amostra.cortes[col_y]
                    x_raw = amostra.df[col_x].to_numpy(dtype=float)
                    x_vals = x_raw[i0:i1+1]
                # garantir mesmo tamanho
                min_len = min(len(x_vals), len(y_vals))
                x_vals = x_vals[:min_len]
                y_vals = y_vals[:min_len]

                ax_plot = axes[j]
                label = amostra.nome if not sobrepor_curvas else f"{amostra.nome} - {col_y}"
                cor_linha = amostra.color
                if sobrepor_curvas and len(amostras) == 1:
                    cor_linha = CORES_TAB10[j % len(CORES_TAB10)]

                line, = ax_plot.plot(
                    x_vals, y_vals,
                    label=label,
                    color=cor_linha,
                    linewidth=amostra.linewidth,
                    linestyle=amostra.linestyle
                )
                self._line_refs[j].append(line)
                self._crosshair_series.append({
                    "ax": ax_plot,
                    "serie_key": f"{j}|{label}",
                    "x": np.asarray(x_vals, dtype=float),
                    "y": np.asarray(y_vals, dtype=float),
                    "label": label,
                    "color": cor_linha,
                })

                # Detecção de picos 
                if self.picos_ativos:
                    novos = self._desenhar_picos(
                        ax_plot, x_vals, y_vals,
                        cor_linha, amostra.nome
                    )
                    self._annots.extend(novos)

        if sobrepor_curvas:
            for j, (ax, col_y) in enumerate(zip(axes, cols_y)):
                cor_eixo = CORES_TAB10[j % len(CORES_TAB10)]
                ylabel = self.ylabels[j] if j < len(self.ylabels) and self.ylabels[j] else col_y
                ax.set_ylabel(ylabel, fontsize=self.fonte_eixo, color=cor_eixo)
                ax.tick_params(axis="y", colors=cor_eixo, labelsize=self.fonte_ticks)
                lado = "left" if ax.yaxis.get_label_position() == "left" else "right"
                ax.spines[lado].set_edgecolor(cor_eixo)
                if lado == "left":
                    ax.spines["right"].set_visible(False)
                else:
                    ax.spines["left"].set_visible(False)
                if j > 0:
                    ax.patch.set_visible(False)
                ax.grid(False)

            ax0 = axes[0]
            ax0.set_xlabel(self.xlabel or col_x, fontsize=self.fonte_eixo)
            ax0.set_title(self.titulo or "Curvas sobrepostas", fontsize=self.fonte_titulo, fontweight="bold")
            ax0.tick_params(axis="x", labelsize=self.fonte_ticks)
            linhas, labels = [], []
            for ax in axes:
                h, l = ax.get_legend_handles_labels()
                linhas.extend(h)
                labels.extend(l)
            ax0.legend(linhas, labels, facecolor="white", edgecolor="#ccc",
                       labelcolor="#111", fontsize=self.fonte_legenda, framealpha=0.9)
            axes_para_formatar = [ax0]
            cols_para_formatar = [cols_y[0]]
        else:
            axes_para_formatar = axes
            cols_para_formatar = cols_y

        if self.xlim_min is not None or self.xlim_max is not None:
            for ax in axes:
                xmin_atual, xmax_atual = ax.get_xlim()
                xmin = self.xlim_min if self.xlim_min is not None else xmin_atual
                xmax = self.xlim_max if self.xlim_max is not None else xmax_atual
                if xmin < xmax:
                    ax.set_xlim(xmin, xmax)

        for j, (ax, col_y) in enumerate(zip(axes_para_formatar, cols_para_formatar)):
            if not sobrepor_curvas:
                ax.set_xlabel(self.xlabel or col_x, fontsize=self.fonte_eixo)
                ylabel = self.ylabels[j] if j < len(self.ylabels) else col_y
                ax.set_ylabel(ylabel, fontsize=self.fonte_eixo)
                titulo = self.titulo if n_y == 1 else f"{col_y}"
                ax.set_title(titulo, fontsize=self.fonte_titulo, fontweight="bold")
                ax.tick_params(axis="both", labelsize=self.fonte_ticks)
                ax.legend(facecolor="white", edgecolor="#ccc",
                          labelcolor="#111", fontsize=self.fonte_legenda, framealpha=0.9)

            # Grid avançado
            ax.grid(False)  # limpa primeiro
            mostrar_algum = self.grid_x or self.grid_y

            if mostrar_algum:
                # Intervalo fixo no eixo X (temperatura)?
                if self.grid_intervalo > 0:
                    x_min, x_max = ax.get_xlim()
                    inicio = math.floor(x_min / self.grid_intervalo) * self.grid_intervalo
                    ticks_x = np.arange(inicio, x_max + self.grid_intervalo, self.grid_intervalo)
                    ax.set_xticks(ticks_x, minor=False)

                gs = {"color": "#ddd", "linestyle": "--", "linewidth": 0.6, "alpha": 0.8}
                if self.grid_x and self.grid_y:
                    ax.grid(True, which="major", **gs)
                elif self.grid_x:
                    ax.xaxis.grid(True, which="major", **gs)
                    ax.yaxis.grid(False)
                elif self.grid_y:
                    ax.yaxis.grid(True, which="major", **gs)
                    ax.xaxis.grid(False)

        for mark in self._fixed_crosshair_marks:
            self._desenhar_crosshair_fixo(mark)

        self.fig.tight_layout(pad=2)
        if sobrepor_curvas and n_y > 1:
            margem_esq = min(0.10 + 0.05 * ((n_y - 1) // 2), 0.28)
            margem_dir = max(0.92 - 0.05 * (n_y // 2), 0.72)
            self.fig.subplots_adjust(left=margem_esq, right=margem_dir)
        self.fig.subplots_adjust(bottom=0.14)
        hint = "Passe o cursor para ler valores; clique para fixar; arraste caixas fixas para reposicionar. Duplo clique em titulo/rotulo para editar"
        if self.picos_ativos and self._annots:
            hint += "  •  Arraste as etiquetas de pico para reposicionar"
        self.fig.text(0.5, 0.02, hint,
                      ha="center", fontsize=8, color="#64748b")

        # re-conectar edição por duplo clique
        if self._cid is not None:
            try: self.canvas.mpl_disconnect(self._cid)
            except Exception: pass

        _root = root_for_popup
        _cols_y = cols_y[:]
        _axes_ref = axes[:]

        def _on_dblclick(event):
            if not event.dblclick:
                return
            try:
                renderer = self.canvas.get_renderer()
            except Exception:
                return
            for j2, ax2 in enumerate(_axes_ref):
                t = ax2.title
                if t.get_window_extent(renderer).contains(event.x, event.y):
                    def _set_t(novo, _ax=ax2, _j=j2):
                        _ax.set_title(novo, fontsize=self.fonte_titulo, fontweight="bold")
                        if _j == 0: self.titulo = novo
                        self.canvas.draw_idle()
                    popup_edit(_root, "Editar título", t.get_text(), _set_t)
                    return
                xl = ax2.xaxis.label
                if xl.get_window_extent(renderer).contains(event.x, event.y):
                    def _set_xl(novo, _ax=ax2):
                        _ax.set_xlabel(novo, fontsize=self.fonte_eixo)
                        self.xlabel = novo
                        self.canvas.draw_idle()
                    popup_edit(_root, "Editar rótulo X", xl.get_text(), _set_xl)
                    return
                yl = ax2.yaxis.label
                if yl.get_window_extent(renderer).contains(event.x, event.y):
                    def _set_yl(novo, _ax=ax2, _j=j2):
                        _ax.set_ylabel(novo, fontsize=self.fonte_eixo)
                        while len(self.ylabels) <= _j:
                            self.ylabels.append("")
                        self.ylabels[_j] = novo
                        self.canvas.draw_idle()
                    popup_edit(_root, "Editar rótulo Y", yl.get_text(), _set_yl)
                    return

        self._cid = self.canvas.mpl_connect("button_press_event", _on_dblclick)
        self._conectar_crosshair()

        # Conecta arrastar se há anotações
        if self._annots:
            self._conectar_drag()
        else:
            self._desconectar_drag()

        self.canvas.draw_idle()

    def salvar_png(self, caminho: str, dpi=300):
        self.fig.savefig(caminho, dpi=dpi, facecolor="white", bbox_inches="tight")

    def salvar_pdf(self, caminho: str):
        self.fig.savefig(caminho, facecolor="white", bbox_inches="tight")


# PAINEL ESQUERDO 

class PainelEsquerdo(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=SURFACE, width=300)
        self.app = app
        self.pack_propagate(False)
        self._build()

    # Estrutura scrollável
    def _build(self):
        # Cabeçalho fixo (não rola)
        header = tk.Frame(self, bg=SURFACE)
        header.pack(fill="x", side="top")
        tk.Label(header, text="H-DMAPlot", font=FONT_TITLE, fg=ACCENT,
                 bg=SURFACE).pack(pady=(14, 2))
        tk.Label(header, text="Análise Termo-Mecânica", font=FONT_SMALL,
                 fg=TEXT_DIM, bg=SURFACE).pack(pady=(0, 6))
        tk.Label(header, text="Autor: Carlos Henrique Amaro da Silva", font=FONT_SMALL,
                 fg=TEXT_DIM, bg=SURFACE).pack(pady=(0, 6))
        sep(header).pack(fill="x", padx=10)

        # Status fixo na base (não rola)
        self.status_var = tk.StringVar(value="Pronto.")
        status_bar = tk.Frame(self, bg=SURFACE)
        status_bar.pack(fill="x", side="bottom")
        sep(status_bar).pack(fill="x", padx=10)
        tk.Label(status_bar, textvariable=self.status_var, font=FONT_SMALL,
                 fg=TEXT_DIM, bg=SURFACE, wraplength=260,
                 justify="left").pack(anchor="w", padx=14, pady=(4, 8))

        # Canvas central que rola
        self._scroll_canvas = tk.Canvas(self, bg=SURFACE, highlightthickness=0)
        self._scrollbar = tk.Scrollbar(self, orient="vertical",
                                       command=self._scroll_canvas.yview)
        self._scroll_canvas.configure(yscrollcommand=self._scrollbar.set)
        self._scrollbar.pack(side="right", fill="y")
        self._scroll_canvas.pack(side="left", fill="both", expand=True)

        # Frame interior que recebe todos os widgets
        self._inner = tk.Frame(self._scroll_canvas, bg=SURFACE)
        self._win_id = self._scroll_canvas.create_window(
            (0, 0), window=self._inner, anchor="nw"
        )

        self._inner.bind("<Configure>", self._on_inner_configure)
        self._scroll_canvas.bind("<Configure>", self._on_canvas_configure)

        # Rolar com mouse
        self._scroll_canvas.bind("<Enter>",
            lambda e: self._scroll_canvas.bind_all("<MouseWheel>", self._on_mousewheel))
        self._scroll_canvas.bind("<Leave>",
            lambda e: self._scroll_canvas.unbind_all("<MouseWheel>"))

        self._populate(self._inner)

    def _on_inner_configure(self, event):
        self._scroll_canvas.configure(
            scrollregion=self._scroll_canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self._scroll_canvas.itemconfig(self._win_id, width=event.width)

    def _on_mousewheel(self, event):
        self._scroll_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    # Conteúdo do painel 
    def _populate(self, p):
        """Popula o frame interior 'p' com todas as seções."""

        # Seção 1: Arquivos
        self._section(p, "1. Arquivos")
        btn(p, "➕  Adicionar .txt", self.app.carregar_arquivos,
            color=ACCENT).pack(fill="x", padx=14, pady=(4, 2))
        btn(p, "🗑  Remover selecionado", self.app.remover_amostra,
            color="#374151").pack(fill="x", padx=14, pady=2)

        list_outer, list_inner = scrolled_frame(p, bg=CARD)
        list_outer.pack(fill="x", padx=10, pady=4)
        list_outer.configure(height=100)
        self.lista_frame = list_inner
        self.lista_rows = []

        sep(p).pack(fill="x", padx=10, pady=6)

        # Seção 2: Eixos 
        self._section(p, "2. Eixos do Gráfico")

        tk.Label(p, text="Eixo X (temperatura):", font=FONT_SMALL,
                 fg=TEXT_DIM, bg=SURFACE).pack(anchor="w", padx=14, pady=(4, 0))
        self.x_var = tk.StringVar()
        self.x_combo = ttk.Combobox(p, textvariable=self.x_var, state="readonly",
                                    font=FONT_SMALL, width=30)
        self.x_combo.pack(fill="x", padx=14, pady=2)
        self.x_combo.bind("<<ComboboxSelected>>", lambda e: self.app.redesenhar())

        tk.Label(p, text="Eixo Y (curvas):", font=FONT_SMALL,
                 fg=TEXT_DIM, bg=SURFACE).pack(anchor="w", padx=14, pady=(6, 0))
        y_select_row = tk.Frame(p, bg=SURFACE)
        y_select_row.pack(fill="x", padx=14, pady=2)
        self.y_var = tk.StringVar()
        self.y_combo = ttk.Combobox(y_select_row, textvariable=self.y_var,
                                    state="readonly", font=FONT_SMALL, width=24)
        self.y_combo.pack(side="left", fill="x", expand=True)
        self.y_combo.bind("<<ComboboxSelected>>", lambda e: self._add_y_col())
        btn(y_select_row, "+", self._add_y_col, color=ACCENT,
            width=3).pack(side="left", padx=(6, 0))

        y_actions = tk.Frame(p, bg=SURFACE)
        y_actions.pack(fill="x", padx=14, pady=(2, 4))
        self.y_count_label = tk.Label(y_actions, text="Nenhuma curva selecionada",
                                      font=FONT_SMALL, fg=TEXT_DIM, bg=SURFACE)
        self.y_count_label.pack(side="left")
        tk.Button(y_actions, text="Limpar", command=self._clear_y_cols,
                  bg=SURFACE, fg=TEXT_DIM, activebackground=SURFACE,
                  activeforeground=TEXT, relief="flat", bd=0,
                  cursor="hand2", font=FONT_SMALL).pack(side="right")

        y_outer, y_inner = scrolled_frame(p, bg=CARD)
        y_outer.pack(fill="x", padx=10, pady=4)
        y_outer.configure(height=150)
        self.y_frame = y_inner
        self.y_checks: dict[str, tk.BooleanVar] = {}

        self.sobrepor_var = tk.BooleanVar(value=False)
        tk.Checkbutton(p, text="Sobrepor curvas no mesmo eixo",
                       variable=self.sobrepor_var,
                       bg=SURFACE, fg=TEXT, font=FONT_SMALL,
                       selectcolor=CARD, activebackground=SURFACE,
                       command=self.app.on_sobreposicao_change).pack(anchor="w", padx=14, pady=(2, 0))

        sep(p).pack(fill="x", padx=10, pady=6)

        # Seção 3: Estilo da Amostra
        self._section(p, "3. Estilo da Amostra")

        tk.Label(p, text="Amostra:", font=FONT_SMALL,
                 fg=TEXT_DIM, bg=SURFACE).pack(anchor="w", padx=14, pady=(4, 0))
        self.amostra_var = tk.StringVar()
        self.amostra_combo = ttk.Combobox(p, textvariable=self.amostra_var,
                                          state="readonly", font=FONT_SMALL, width=30)
        self.amostra_combo.pack(fill="x", padx=14, pady=2)
        self.amostra_combo.bind("<<ComboboxSelected>>",
                                lambda e: self.app.on_amostra_select(self.amostra_var.get()))

        est_row = tk.Frame(p, bg=SURFACE)
        est_row.pack(fill="x", padx=14, pady=4)
        tk.Label(est_row, text="Tipo:", font=FONT_SMALL, fg=TEXT_DIM,
                 bg=SURFACE).pack(side="left")
        self.estilo_var = tk.StringVar(value="Sólida")
        estilo_combo = ttk.Combobox(est_row, textvariable=self.estilo_var,
                                    values=list(ESTILOS_LINHA.keys()),
                                    state="readonly", font=FONT_SMALL, width=12)
        estilo_combo.pack(side="left", padx=6)
        estilo_combo.bind("<<ComboboxSelected>>",
                          lambda e: self.app.on_estilo_change(self.estilo_var.get()))

        lw_row = tk.Frame(p, bg=SURFACE)
        lw_row.pack(fill="x", padx=14, pady=2)
        tk.Label(lw_row, text="Espessura:", font=FONT_SMALL,
                 fg=TEXT_DIM, bg=SURFACE).pack(side="left")
        self.lw_var = tk.DoubleVar(value=1.8)
        self.lw_label = tk.Label(lw_row, text="1.8 pt", font=FONT_SMALL,
                                 fg=TEXT, bg=SURFACE, width=6)
        self.lw_label.pack(side="right")
        tk.Scale(lw_row, from_=0.5, to=5.0, resolution=0.1,
                 orient="horizontal", variable=self.lw_var,
                 bg=SURFACE, fg=TEXT, troughcolor=CARD,
                 highlightthickness=0, sliderrelief="flat",
                 command=self.app.on_lw_change, length=100).pack(side="left", padx=4)

        cor_row = tk.Frame(p, bg=SURFACE)
        cor_row.pack(fill="x", padx=14, pady=4)
        tk.Label(cor_row, text="Cor:", font=FONT_SMALL,
                 fg=TEXT_DIM, bg=SURFACE).pack(side="left")
        self.cor_preview = tk.Label(cor_row, bg=CORES_TAB10[0],
                                    width=4, relief="flat", cursor="hand2")
        self.cor_preview.pack(side="left", padx=6)
        self.cor_preview.bind("<Button-1>", lambda e: self.app.escolher_cor())

        sep(p).pack(fill="x", padx=10, pady=6)

        # Seção 4: Suavização
        self._section(p, "4. Suavização de Curvas")

        # Seletor de método
        met_row = tk.Frame(p, bg=SURFACE)
        met_row.pack(fill="x", padx=14, pady=(4, 2))
        tk.Label(met_row, text="Método:", font=FONT_SMALL,
                 fg=TEXT_DIM, bg=SURFACE).pack(side="left")
        self.suav_metodo_var = tk.StringVar(value="Savitzky-Golay")
        met_combo = ttk.Combobox(met_row, textvariable=self.suav_metodo_var,
                                 values=["Savitzky-Golay", "Média Móvel"],
                                 state="readonly", font=FONT_SMALL, width=16)
        met_combo.pack(side="left", padx=6)

        # Janela
        jan_row = tk.Frame(p, bg=SURFACE)
        jan_row.pack(fill="x", padx=14, pady=2)
        tk.Label(jan_row, text="Janela (pts):", font=FONT_SMALL,
                 fg=TEXT_DIM, bg=SURFACE).pack(side="left")
        self.suav_janela_var = tk.IntVar(value=11)
        self.suav_janela_label = tk.Label(jan_row, text="11", font=FONT_SMALL,
                                          fg=TEXT, bg=SURFACE, width=4)
        self.suav_janela_label.pack(side="right")
        tk.Scale(jan_row, from_=3, to=101, resolution=2,
                 orient="horizontal", variable=self.suav_janela_var,
                 bg=SURFACE, fg=TEXT, troughcolor=CARD,
                 highlightthickness=0, sliderrelief="flat",
                 command=lambda v: self.suav_janela_label.config(text=str(int(float(v)))),
                 length=110).pack(side="left", padx=4)

        # Grau polinomial (SavGol)
        poly_row = tk.Frame(p, bg=SURFACE)
        poly_row.pack(fill="x", padx=14, pady=2)
        tk.Label(poly_row, text="Grau poly (SG):", font=FONT_SMALL,
                 fg=TEXT_DIM, bg=SURFACE).pack(side="left")
        self.suav_poly_var = tk.IntVar(value=3)
        self.suav_poly_label = tk.Label(poly_row, text="3", font=FONT_SMALL,
                                        fg=TEXT, bg=SURFACE, width=4)
        self.suav_poly_label.pack(side="right")
        tk.Scale(poly_row, from_=1, to=9, resolution=1,
                 orient="horizontal", variable=self.suav_poly_var,
                 bg=SURFACE, fg=TEXT, troughcolor=CARD,
                 highlightthickness=0, sliderrelief="flat",
                 command=lambda v: self.suav_poly_label.config(text=str(int(float(v)))),
                 length=110).pack(side="left", padx=4)

        # Botões de ação
        btn(p, "〜  Aplicar suavização", self.app.aplicar_suavizacao,
            color=WARNING).pack(fill="x", padx=14, pady=(6, 2))
        btn(p, "✕  Remover suavização", self.app.remover_suavizacao,
            color="#374151").pack(fill="x", padx=14, pady=2)

        sep(p).pack(fill="x", padx=10, pady=6)

        # Seção 5: Tratamento de Dados
        self._section(p, "5. Tratamento de Dados")
        btn(p, "✂  Corte de intervalo", self.app.abrir_corte,
            color="#374151").pack(fill="x", padx=14, pady=2)
        btn(p, "↕  Offset numérico", self.app.abrir_offset,
            color="#374151").pack(fill="x", padx=14, pady=2)
        btn(p, "Limpar marcações fixas", self.app.limpar_marcacoes_fixas,
            color="#374151").pack(fill="x", padx=14, pady=2)
        btn(p, "↺  Resetar todos os tratamentos", self.app.resetar_tratamentos,
            color="#374151").pack(fill="x", padx=14, pady=2)

        sep(p).pack(fill="x", padx=10, pady=6)

        # Seção 6: Grid e Fontes 
        self._section(p, "6. Grade do Gráfico")

        # Checkboxes X e Y lado a lado
        grid_chk_row = tk.Frame(p, bg=SURFACE)
        grid_chk_row.pack(fill="x", padx=14, pady=(4, 2))
        self.grid_x_var = tk.BooleanVar(value=True)
        self.grid_y_var = tk.BooleanVar(value=True)
        tk.Checkbutton(grid_chk_row, text="Linhas X (vert.)", variable=self.grid_x_var,
                       bg=SURFACE, fg=TEXT, font=FONT_SMALL, selectcolor=CARD,
                       activebackground=SURFACE,
                       command=self.app.on_grid_change).pack(side="left", padx=(0, 10))
        tk.Checkbutton(grid_chk_row, text="Linhas Y (horiz.)", variable=self.grid_y_var,
                       bg=SURFACE, fg=TEXT, font=FONT_SMALL, selectcolor=CARD,
                       activebackground=SURFACE,
                       command=self.app.on_grid_change).pack(side="left")

        # Intervalo do grid
        INTERVALOS_GRID = ["Automático", "1", "2", "5", "10", "20", "25", "50", "100"]
        intv_row = tk.Frame(p, bg=SURFACE)
        intv_row.pack(fill="x", padx=14, pady=(4, 2))
        tk.Label(intv_row, text="Intervalo X:", font=FONT_SMALL,
                 fg=TEXT_DIM, bg=SURFACE).pack(side="left")
        self.grid_intervalo_var = tk.StringVar(value="Automático")
        intv_combo = ttk.Combobox(intv_row, textvariable=self.grid_intervalo_var,
                                  values=INTERVALOS_GRID, state="readonly",
                                  font=FONT_SMALL, width=10)
        intv_combo.pack(side="left", padx=6)
        intv_combo.bind("<<ComboboxSelected>>", lambda e: self.app.on_grid_change())

        xlim_row = tk.Frame(p, bg=SURFACE)
        xlim_row.pack(fill="x", padx=14, pady=(4, 2))
        tk.Label(xlim_row, text="Limite X:", font=FONT_SMALL,
                 fg=TEXT_DIM, bg=SURFACE).pack(side="left")
        self.xlim_min_var = tk.StringVar(value="")
        self.xlim_max_var = tk.StringVar(value="")
        xlim_min_entry = tk.Entry(xlim_row, textvariable=self.xlim_min_var,
                                  font=FONT_SMALL, width=7, bg=CARD, fg=TEXT,
                                  insertbackground=TEXT, relief="flat")
        xlim_min_entry.pack(side="left", padx=(6, 2), ipady=2)
        tk.Label(xlim_row, text="a", font=FONT_SMALL,
                 fg=TEXT_DIM, bg=SURFACE).pack(side="left")
        xlim_max_entry = tk.Entry(xlim_row, textvariable=self.xlim_max_var,
                                  font=FONT_SMALL, width=7, bg=CARD, fg=TEXT,
                                  insertbackground=TEXT, relief="flat")
        xlim_max_entry.pack(side="left", padx=(2, 0), ipady=2)
        xlim_min_entry.bind("<Return>", lambda e: self.app.on_grid_change())
        xlim_max_entry.bind("<Return>", lambda e: self.app.on_grid_change())

        xlim_btn_row = tk.Frame(p, bg=SURFACE)
        xlim_btn_row.pack(fill="x", padx=14, pady=(0, 2))
        tk.Button(xlim_btn_row, text="Aplicar limites", command=self.app.on_grid_change,
                  bg="#374151", fg=TEXT, activebackground="#4b5563",
                  activeforeground=TEXT, relief="flat", bd=0,
                  cursor="hand2", font=FONT_SMALL).pack(side="left", padx=(0, 4))
        tk.Button(xlim_btn_row, text="Auto", command=self.limpar_xlim,
                  bg=SURFACE, fg=TEXT_DIM, activebackground=SURFACE,
                  activeforeground=TEXT, relief="flat", bd=0,
                  cursor="hand2", font=FONT_SMALL).pack(side="left")

        sep(p).pack(fill="x", padx=10, pady=6)

        # Seção 6b: Fontes
        self._section(p, "6b. Tamanho de Fontes")

        _FONT_VALS = list(range(6, 22))

        def _font_row(parent, label, attr_name, default):
            row = tk.Frame(parent, bg=SURFACE)
            row.pack(fill="x", padx=14, pady=2)
            tk.Label(row, text=label, font=FONT_SMALL, fg=TEXT_DIM,
                     bg=SURFACE, width=12, anchor="w").pack(side="left")
            var = tk.IntVar(value=default)
            setattr(self, attr_name, var)
            lbl_val = tk.Label(row, text=str(default), font=FONT_SMALL,
                               fg=TEXT, bg=SURFACE, width=3)
            lbl_val.pack(side="right")
            tk.Scale(row, from_=6, to=20, resolution=1, orient="horizontal",
                     variable=var, bg=SURFACE, fg=TEXT, troughcolor=CARD,
                     highlightthickness=0, sliderrelief="flat",
                     command=lambda v, lv=lbl_val: [
                         lv.config(text=str(int(float(v)))),
                         self.app.on_fonte_change()
                     ],
                     length=90).pack(side="left", padx=4)

        _font_row(p, "Título:",    "fonte_titulo_var",  11)
        _font_row(p, "Rót. eixo:", "fonte_eixo_var",    10)
        _font_row(p, "Ticks:",     "fonte_ticks_var",    9)
        _font_row(p, "Legenda:",   "fonte_legenda_var",  9)
        _font_row(p, "Picos:",     "fonte_picos_var",    8)

        sep(p).pack(fill="x", padx=10, pady=6)

        # Seção 7: Detecção de Picos 
        self._section(p, "7. Detecção de Picos")

        # Ligar/desligar
        self.picos_var = tk.BooleanVar(value=False)
        tk.Checkbutton(p, text="Mostrar picos nas curvas", variable=self.picos_var,
                       bg=SURFACE, fg=TEXT, font=FONT_SMALL,
                       selectcolor=CARD, activebackground=SURFACE,
                       command=self.app.on_picos_toggle).pack(anchor="w", padx=14)

        # Proeminência mínima
        prom_row = tk.Frame(p, bg=SURFACE)
        prom_row.pack(fill="x", padx=14, pady=(4, 2))
        tk.Label(prom_row, text="Proeminência (%):", font=FONT_SMALL,
                 fg=TEXT_DIM, bg=SURFACE).pack(side="left")
        self.picos_prom_var = tk.IntVar(value=5)
        self.picos_prom_label = tk.Label(prom_row, text="5%", font=FONT_SMALL,
                                         fg=TEXT, bg=SURFACE, width=5)
        self.picos_prom_label.pack(side="right")
        tk.Scale(prom_row, from_=1, to=50, resolution=1,
                 orient="horizontal", variable=self.picos_prom_var,
                 bg=SURFACE, fg=TEXT, troughcolor=CARD,
                 highlightthickness=0, sliderrelief="flat",
                 command=lambda v: [
                     self.picos_prom_label.config(text=f"{int(float(v))}%"),
                     self.app.on_picos_cfg_change()
                 ],
                 length=100).pack(side="left", padx=4)

        # Distância mínima entre picos
        dist_row = tk.Frame(p, bg=SURFACE)
        dist_row.pack(fill="x", padx=14, pady=2)
        tk.Label(dist_row, text="Distância (pts):", font=FONT_SMALL,
                 fg=TEXT_DIM, bg=SURFACE).pack(side="left")
        self.picos_dist_var = tk.IntVar(value=10)
        self.picos_dist_label = tk.Label(dist_row, text="10", font=FONT_SMALL,
                                         fg=TEXT, bg=SURFACE, width=5)
        self.picos_dist_label.pack(side="right")
        tk.Scale(dist_row, from_=1, to=100, resolution=1,
                 orient="horizontal", variable=self.picos_dist_var,
                 bg=SURFACE, fg=TEXT, troughcolor=CARD,
                 highlightthickness=0, sliderrelief="flat",
                 command=lambda v: [
                     self.picos_dist_label.config(text=str(int(float(v)))),
                     self.app.on_picos_cfg_change()
                 ],
                 length=100).pack(side="left", padx=4)

        # Máximo de picos por curva
        max_row = tk.Frame(p, bg=SURFACE)
        max_row.pack(fill="x", padx=14, pady=2)
        tk.Label(max_row, text="Máx. picos/curva:", font=FONT_SMALL,
                 fg=TEXT_DIM, bg=SURFACE).pack(side="left")
        self.picos_max_var = tk.IntVar(value=5)
        self.picos_max_label = tk.Label(max_row, text="5", font=FONT_SMALL,
                                        fg=TEXT, bg=SURFACE, width=5)
        self.picos_max_label.pack(side="right")
        tk.Scale(max_row, from_=1, to=20, resolution=1,
                 orient="horizontal", variable=self.picos_max_var,
                 bg=SURFACE, fg=TEXT, troughcolor=CARD,
                 highlightthickness=0, sliderrelief="flat",
                 command=lambda v: [
                     self.picos_max_label.config(text=str(int(float(v)))),
                     self.app.on_picos_cfg_change()
                 ],
                 length=100).pack(side="left", padx=4)

        sep(p).pack(fill="x", padx=10, pady=6)

        # Seção 8: Exportar 
        self._section(p, "8. Exportar")
        btn(p, "📊  Exportar Excel (.xlsx)", self.app.exportar_excel,
            color=SUCCESS).pack(fill="x", padx=14, pady=2)
        btn(p, "🖼  Exportar PNG", self.app.exportar_png,
            color=ACCENT2).pack(fill="x", padx=14, pady=2)
        btn(p, "📄  Exportar PDF", self.app.exportar_pdf,
            color="#dc2626").pack(fill="x", padx=14, pady=(2, 12))

    def _section(self, parent, titulo):
        f = tk.Frame(parent, bg=SURFACE)
        f.pack(fill="x", padx=10, pady=(6, 2))
        tk.Label(f, text=titulo, font=("Segoe UI", 9, "bold"),
                 fg=ACCENT, bg=SURFACE).pack(anchor="w")

    # Atualização da lista de arquivos 
    def atualizar_lista(self, amostras: list[Amostra], selecionado_idx: int | None):
        for w in self.lista_frame.winfo_children():
            w.destroy()
        self.lista_rows = []

        if not amostras:
            tk.Label(self.lista_frame, text="Nenhum arquivo carregado.",
                     font=FONT_SMALL, fg=TEXT_DIM, bg=CARD).pack(anchor="w", padx=6, pady=4)
            return

        for i, am in enumerate(amostras):
            selected = (i == selecionado_idx)
            bg = ACCENT2 if selected else CARD
            fg = "white" if selected else TEXT

            row = tk.Frame(self.lista_frame, bg=bg, cursor="hand2")
            row.pack(fill="x", pady=1)
            self.lista_rows.append(row)

            dot = tk.Label(row, text="●", font=FONT_SMALL, fg=am.color, bg=bg)
            dot.pack(side="left", padx=4)
            nome_lbl = tk.Label(row, text=am.nome, font=FONT_MONO,
                                fg=fg, bg=bg, anchor="w")
            nome_lbl.pack(side="left", fill="x", expand=True, padx=2)

            for w in (row, dot, nome_lbl):
                w.bind("<Button-1>", lambda e, idx=i: self.app.selecionar_amostra(idx))

    # Atualização das colunas nos combos 
    def atualizar_colunas(self, colunas: list[str]):
        # Eixo X - mostra rotulo tecnico no combo (curto, cabe bem)
        self.x_combo["values"] = colunas
        if not self.x_var.get() and colunas:
            for c in colunas:
                if "ts" in c.lower() or "temp" in c.lower():
                    self.x_var.set(c)
                    break
            else:
                self.x_var.set(colunas[0] if colunas else "")

        # Eixo Y - usa o mesmo padrao de combo do X, mantendo multipla selecao.
        prev_checked = {c for c, v in self.y_checks.items() if v.get()}
        self.y_combo["values"] = colunas
        if colunas and (not self.y_var.get() or self.y_var.get() not in colunas):
            preferida = next((c for c in colunas if "tan" in c.lower()), None)
            self.y_var.set(preferida or colunas[0])
        elif not colunas:
            self.y_var.set("")

        self.y_checks = {
            c: tk.BooleanVar(value=(c in prev_checked))
            for c in colunas
        }
        self._render_y_selected()

    def _add_y_col(self):
        col = self.y_var.get()
        if not col or col not in self.y_checks:
            return
        if not self.y_checks[col].get():
            self.y_checks[col].set(True)
            self._render_y_selected()
            self.app.redesenhar()

    def _remove_y_col(self, col: str):
        if col in self.y_checks and self.y_checks[col].get():
            self.y_checks[col].set(False)
            self._render_y_selected()
            self.app.redesenhar()

    def _clear_y_cols(self):
        changed = False
        for var in self.y_checks.values():
            if var.get():
                var.set(False)
                changed = True
        if changed:
            self._render_y_selected()
            self.app.redesenhar()

    def _render_y_selected(self):
        for w in self.y_frame.winfo_children():
            w.destroy()

        selected = [c for c, v in self.y_checks.items() if v.get()]
        if not selected:
            tk.Label(self.y_frame, text="Adicione uma coluna para plotar no eixo Y.",
                     font=FONT_SMALL, fg=TEXT_DIM, bg=CARD,
                     wraplength=220, justify="left").pack(anchor="w", padx=8, pady=8)
            self._sync_y_count(0)
            return

        for c in selected:
            row = tk.Frame(self.y_frame, bg=CARD)
            row.pack(fill="x", pady=2, padx=2)

            txt_frame = tk.Frame(row, bg=CARD)
            txt_frame.pack(side="left", fill="x", expand=True, padx=(6, 4), pady=4)

            nome_col = c.split(" [")[0].strip()
            desc = DESCRICOES_COLUNAS.get(nome_col) or DESCRICOES_COLUNAS.get(nome_col.lower(), "")
            if desc:
                tk.Label(txt_frame, text=desc, font=("Segoe UI", 8, "bold"),
                         fg=TEXT, bg=CARD, anchor="w", wraplength=185,
                         justify="left").pack(anchor="w")

            tk.Label(txt_frame, text=c, font=FONT_MONO,
                     fg=TEXT_DIM, bg=CARD, anchor="w").pack(anchor="w")

            tk.Button(row, text="x", command=lambda col=c: self._remove_y_col(col),
                      bg=CARD, fg=TEXT_DIM, activebackground=CARD,
                      activeforeground=ERROR, relief="flat", bd=0,
                      cursor="hand2", font=("Segoe UI", 10, "bold"),
                      width=2).pack(side="right", padx=(0, 5), pady=4)

        self._sync_y_count(len(selected))

    def _sync_y_count(self, total: int):
        if total == 0:
            texto = "Nenhuma curva selecionada"
        elif total == 1:
            texto = "1 curva selecionada"
        else:
            texto = f"{total} curvas selecionadas"
        self.y_count_label.config(text=texto)

    def atualizar_amostra_combo(self, amostras: list[Amostra]):
        nomes = [a.nome for a in amostras]
        self.amostra_combo["values"] = nomes
        if nomes and not self.amostra_var.get():
            self.amostra_var.set(nomes[0])

    def get_col_x(self) -> str:
        return self.x_var.get()

    def get_cols_y(self) -> list[str]:
        return [c for c, v in self.y_checks.items() if v.get()]

    def get_sobreposicao_cfg(self) -> bool:
        return self.sobrepor_var.get()

    def limpar_xlim(self):
        self.xlim_min_var.set("")
        self.xlim_max_var.set("")
        self.app.on_grid_change()

    def sincronizar_estilo(self, amostra: Amostra):
        nome_estilo = next((k for k, v in ESTILOS_LINHA.items()
                            if v == amostra.linestyle), "Sólida")
        self.estilo_var.set(nome_estilo)
        self.lw_var.set(amostra.linewidth)
        self.lw_label.config(text=f"{amostra.linewidth:.1f} pt")
        self.cor_preview.config(bg=amostra.color)

    def get_suavizacao_cfg(self) -> dict:
        metodo_label = self.suav_metodo_var.get()
        metodo = "savgol" if "Savitzky" in metodo_label else "media_movel"
        janela = int(self.suav_janela_var.get())
        if janela % 2 == 0:
            janela += 1
        return {
            "metodo": metodo,
            "janela": janela,
            "poly": int(self.suav_poly_var.get()),
        }

    def get_grid_cfg(self) -> dict:
        raw = self.grid_intervalo_var.get()
        try:
            intervalo = float(raw.replace(",", "."))
        except ValueError:
            intervalo = 0.0

        def _parse_limite(var, nome):
            txt = var.get().strip().replace(",", ".")
            if not txt:
                return None
            try:
                return float(txt)
            except ValueError:
                messagebox.showerror("Limite X", f"Digite um valor numerico valido para {nome}.")
                return None

        xlim_min = _parse_limite(self.xlim_min_var, "X min")
        xlim_max = _parse_limite(self.xlim_max_var, "X max")
        if xlim_min is not None and xlim_max is not None and xlim_min >= xlim_max:
            messagebox.showerror("Limite X", "O limite minimo de X deve ser menor que o limite maximo.")
            xlim_min = None
            xlim_max = None

        return {
            "grid_x":   self.grid_x_var.get(),
            "grid_y":   self.grid_y_var.get(),
            "intervalo": intervalo,
            "xlim_min":  xlim_min,
            "xlim_max":  xlim_max,
        }

    def get_fontes_cfg(self) -> dict:
        return {
            "titulo":  int(self.fonte_titulo_var.get()),
            "eixo":    int(self.fonte_eixo_var.get()),
            "ticks":   int(self.fonte_ticks_var.get()),
            "legenda": int(self.fonte_legenda_var.get()),
            "picos":   int(self.fonte_picos_var.get()),
        }

    def get_picos_cfg(self) -> dict:
        return {
            "ativo":      self.picos_var.get(),
            "prominence": int(self.picos_prom_var.get()) / 100.0,
            "distancia":  int(self.picos_dist_var.get()),
            "max_picos":  int(self.picos_max_var.get()),
        }

    def status(self, msg: str):
        self.status_var.set(msg)


# DIALOGS DE TRATAMENTO

class DialogCorte(tk.Toplevel):
    """Define intervalo de corte por valor real da coluna (ex: temperatura, tempo)."""
    def __init__(self, parent, amostra: Amostra, col: str, callback):
        super().__init__(parent)
        self.title(f"Corte — {amostra.nome} / {col}")
        self.configure(bg="white")
        self.resizable(False, False)
        self.grab_set()
        self.attributes("-topmost", True)

        vals = amostra.df[col].to_numpy(dtype=float)
        n = len(vals)
        v_min = float(np.nanmin(vals))
        v_max = float(np.nanmax(vals))

        # Se já havia corte anterior, pré-preenche com os valores reais correspondentes
        prev = amostra.cortes.get(col)
        if prev:
            val_ini = float(vals[prev[0]])
            val_fim = float(vals[prev[1]])
        else:
            val_ini, val_fim = v_min, v_max

        # Informações da coluna
        info = tk.Frame(self, bg="white")
        info.pack(padx=20, pady=(14, 6))
        tk.Label(info, text=f"Coluna:  {col}", font=("Segoe UI", 9, "bold"),
                 bg="white", fg="#333").grid(row=0, column=0, columnspan=2, sticky="w")
        tk.Label(info, text=f"Intervalo disponível:  {v_min:.4g}  →  {v_max:.4g}",
                 font=("Segoe UI", 9), bg="white", fg="#666").grid(row=1, column=0, columnspan=2, sticky="w", pady=(2, 0))
        tk.Label(info, text=f"Total de pontos: {n}",
                 font=("Segoe UI", 9), bg="white", fg="#666").grid(row=2, column=0, columnspan=2, sticky="w")

        sep_line = tk.Frame(self, bg="#e5e7eb", height=1)
        sep_line.pack(fill="x", padx=20, pady=8)

        def _entry_row(label_text, default_val):
            row = tk.Frame(self, bg="white")
            row.pack(padx=20, pady=4, fill="x")
            tk.Label(row, text=label_text, bg="white", font=("Segoe UI", 9),
                     width=14, anchor="w").pack(side="left")
            var = tk.StringVar(value=f"{default_val:.6g}")
            e = tk.Entry(row, textvariable=var, font=("Segoe UI", 10),
                         bg="white", fg="#111", relief="flat",
                         highlightthickness=1, highlightcolor=ACCENT,
                         highlightbackground="#ccc", width=14)
            e.pack(side="left", ipady=4, padx=(4, 0))
            return var

        self.ini_var = _entry_row("Valor inicial:", val_ini)
        self.fim_var = _entry_row("Valor final:", val_fim)

        tk.Label(self, text="Os valores serão convertidos para os índices mais próximos.",
                 font=("Segoe UI", 8), bg="white", fg="#aaa").pack(padx=20, pady=(4, 2))

        bf = tk.Frame(self, bg="white")
        bf.pack(pady=(6, 14))
        tk.Button(bf, text="Cancelar", command=self.destroy,
                  bg="#f1f5f9", fg="#555", font=("Segoe UI", 9),
                  relief="flat", padx=10, pady=4).pack(side="left", padx=4)
        tk.Button(bf, text="Aplicar", font=("Segoe UI", 9, "bold"),
                  bg=ACCENT, fg="white", relief="flat", padx=10, pady=4,
                  command=lambda: self._apply(callback, vals, n)).pack(side="left")

    def _apply(self, callback, vals, n):
        try:
            v_ini = float(self.ini_var.get().replace(",", "."))
            v_fim = float(self.fim_var.get().replace(",", "."))
        except ValueError:
            messagebox.showerror("Corte", "Digite valores numéricos válidos.")
            return
        # Converte valor real → índice mais próximo
        i0 = int(np.searchsorted(vals, v_ini, side="left"))
        i1 = int(np.searchsorted(vals, v_fim, side="right")) - 1
        i0 = max(0, min(i0, n - 1))
        i1 = max(i0, min(i1, n - 1))
        callback(i0, i1)
        self.destroy()


class DialogOffset(tk.Toplevel):
    """Define offset numérico para uma coluna de uma amostra."""
    def __init__(self, parent, amostra: Amostra, col: str, callback):
        super().__init__(parent)
        self.title(f"Offset — {amostra.nome} / {col}")
        self.configure(bg="white")
        self.resizable(False, False)
        self.grab_set()
        self.attributes("-topmost", True)

        prev = amostra.offsets.get(col, 0.0)

        tk.Label(self, text=f"Offset para '{col}':",
                 font=("Segoe UI", 9), bg="white", fg="#555").pack(padx=20, pady=(14, 4))

        self.val_var = tk.DoubleVar(value=prev)
        tk.Entry(self, textvariable=self.val_var, font=("Segoe UI", 11),
                 bg="white", fg="#111", relief="flat", highlightthickness=1,
                 highlightcolor=ACCENT, highlightbackground="#ccc",
                 width=16).pack(padx=20, pady=6, ipady=4)

        bf = tk.Frame(self, bg="white")
        bf.pack(pady=(4, 14))
        tk.Button(bf, text="Cancelar", command=self.destroy,
                  bg="#f1f5f9", fg="#555", font=("Segoe UI", 9),
                  relief="flat", padx=10, pady=4).pack(side="left", padx=4)
        tk.Button(bf, text="Aplicar", font=("Segoe UI", 9, "bold"),
                  bg=ACCENT, fg="white", relief="flat", padx=10, pady=4,
                  command=lambda: self._apply(callback)).pack(side="left")

    def _apply(self, callback):
        callback(self.val_var.get())
        self.destroy()


class DialogSelecionarAmostraColuna(tk.Toplevel):
    """Permite ao usuário escolher qual amostra e coluna aplicar tratamento."""
    def __init__(self, parent, amostras: list[Amostra], titulo: str, callback):
        super().__init__(parent)
        self.title(titulo)
        self.configure(bg="white")
        self.resizable(False, False)
        self.grab_set()
        self.attributes("-topmost", True)

        tk.Label(self, text="Amostra:", font=("Segoe UI", 9, "bold"),
                 bg="white", fg="#555").pack(anchor="w", padx=20, pady=(14, 2))
        self.am_var = tk.StringVar()
        am_combo = ttk.Combobox(self, textvariable=self.am_var, state="readonly",
                                values=[a.nome for a in amostras], font=("Segoe UI", 9), width=28)
        am_combo.pack(padx=20, pady=2)
        if amostras:
            am_combo.current(0)

        tk.Label(self, text="Coluna:", font=("Segoe UI", 9, "bold"),
                 bg="white", fg="#555").pack(anchor="w", padx=20, pady=(10, 2))
        self.col_var = tk.StringVar()
        colunas = amostras[0].colunas if amostras else []
        col_combo = ttk.Combobox(self, textvariable=self.col_var, state="readonly",
                                 values=colunas, font=("Segoe UI", 9), width=28)
        col_combo.pack(padx=20, pady=2)
        if colunas:
            col_combo.current(0)

        def _update_cols(event):
            nome = self.am_var.get()
            am = next((a for a in amostras if a.nome == nome), None)
            if am:
                col_combo["values"] = am.colunas
                col_combo.current(0)

        am_combo.bind("<<ComboboxSelected>>", _update_cols)

        bf = tk.Frame(self, bg="white")
        bf.pack(pady=(14, 14))
        tk.Button(bf, text="Cancelar", command=self.destroy,
                  bg="#f1f5f9", fg="#555", font=("Segoe UI", 9),
                  relief="flat", padx=10, pady=4).pack(side="left", padx=4)
        tk.Button(bf, text="Continuar →", font=("Segoe UI", 9, "bold"),
                  bg=ACCENT, fg="white", relief="flat", padx=10, pady=4,
                  command=lambda: self._ok(amostras, callback)).pack(side="left")

    def _ok(self, amostras, callback):
        nome = self.am_var.get()
        col  = self.col_var.get()
        am   = next((a for a in amostras if a.nome == nome), None)
        self.destroy()
        if am and col:
            callback(am, col)


# APLICAÇÃO PRINCIPAL 

class App:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("H-DMAPlot — Análise Termo-Mecânica v1.3")
        try:
            self.root.iconbitmap(resource_path("DMAPlot.ico"))
        except Exception:
            pass  # ícone não encontrado — continua sem ele
        self.root.configure(bg=BG)
        self.root.minsize(1100, 680)
        try:
            self.root.state("zoomed")
        except Exception:
            self.root.geometry("1200x750")

        self.amostras: list[Amostra] = []
        self.selecionado_idx: int | None = None
        self._redraw_after_id = None

        # Layout principal
        pane = tk.PanedWindow(self.root, orient="horizontal",
                              bg=BG, sashwidth=5, sashrelief="flat")
        pane.pack(fill="both", expand=True)

        self.painel = PainelEsquerdo(pane, self)
        pane.add(self.painel, minsize=300)

        right = tk.Frame(pane, bg="white")
        pane.add(right, minsize=600)
        self.canvas_plot = PlotCanvas(right)

        self.root.mainloop()

    # Arquivo

    def carregar_arquivos(self):
        paths = filedialog.askopenfilenames(
            title="Selecionar arquivos .txt",
            filetypes=[("Arquivos TXT", "*.txt"), ("Todos", "*.*")]
        )
        if not paths:
            return
        carregados = 0
        erros = []
        for path in paths:
            nome = os.path.splitext(os.path.basename(path))[0]
            # evitar duplicata
            if any(a.caminho == path for a in self.amostras):
                continue
            df = _parse_txt(path)
            if df is None or df.empty:
                erros.append(nome)
                continue
            self.amostras.append(Amostra(nome, path, df))
            carregados += 1

        if erros:
            messagebox.showwarning("H-DMAPlot",
                f"Não foi possível ler {len(erros)} arquivo(s):\n" + "\n".join(erros))

        if carregados:
            self._atualizar_tudo()
            self.painel.status(f"{carregados} arquivo(s) carregado(s).")

    def remover_amostra(self):
        if self.selecionado_idx is None or not self.amostras:
            return
        nome = self.amostras[self.selecionado_idx].nome
        self.amostras.pop(self.selecionado_idx)
        self.selecionado_idx = None
        self._atualizar_tudo()
        self.painel.status(f"Amostra '{nome}' removida.")

    def _atualizar_tudo(self):
        self.painel.atualizar_lista(self.amostras, self.selecionado_idx)
        self.painel.atualizar_amostra_combo(self.amostras)
        # colunas em comum
        if self.amostras:
            cols = list(self.amostras[0].colunas)
            demais = [set(am.colunas) for am in self.amostras[1:]]
            if demais:
                cols = [c for c in cols if all(c in s for s in demais)]
            self.painel.atualizar_colunas(cols)
        self.redesenhar()

    # Seleção de amostra

    def selecionar_amostra(self, idx: int):
        self.selecionado_idx = idx
        self.painel.atualizar_lista(self.amostras, idx)
        if idx < len(self.amostras):
            self.painel.amostra_var.set(self.amostras[idx].nome)
            self.painel.sincronizar_estilo(self.amostras[idx])

    def on_amostra_select(self, nome: str):
        idx = next((i for i, a in enumerate(self.amostras) if a.nome == nome), None)
        if idx is not None:
            self.selecionar_amostra(idx)

    def _amostra_ativa(self) -> Amostra | None:
        if self.selecionado_idx is not None and self.selecionado_idx < len(self.amostras):
            return self.amostras[self.selecionado_idx]
        nome = self.painel.amostra_var.get()
        return next((a for a in self.amostras if a.nome == nome), None)

    # Estilo

    def on_estilo_change(self, nome: str):
        am = self._amostra_ativa()
        if am:
            am.linestyle = ESTILOS_LINHA[nome]
            self.redesenhar()

    def on_lw_change(self, val):
        am = self._amostra_ativa()
        if am:
            am.linewidth = float(val)
            self.painel.lw_label.config(text=f"{float(val):.1f} pt")
            self.redesenhar()

    def escolher_cor(self):
        am = self._amostra_ativa()
        if not am:
            return
        resultado = colorchooser.askcolor(color=am.color,
                                          title=f"Cor — {am.nome}",
                                          parent=self.root)
        if resultado and resultado[1]:
            am.color = resultado[1]
            self.painel.cor_preview.config(bg=am.color)
            self.painel.atualizar_lista(self.amostras, self.selecionado_idx)
            self.redesenhar()

    # Redesenhar

    def redesenhar(self, delay_ms: int = 80):
        if self._redraw_after_id is not None:
            try:
                self.root.after_cancel(self._redraw_after_id)
            except Exception:
                pass
        self._redraw_after_id = self.root.after(delay_ms, self._redesenhar_agora)

    def _redesenhar_agora(self):
        self._redraw_after_id = None
        col_x  = self.painel.get_col_x()
        cols_y = self.painel.get_cols_y()
        # aplica configurações de grid e fontes no canvas antes de redesenhar
        self._aplicar_cfg_grid()
        self._aplicar_cfg_fontes()
        self.canvas_plot.redesenhar(
            self.amostras, col_x, cols_y,
            mostrar_grid=None,   # grid agora gerenciado internamente
            root_for_popup=self.root,
            sobrepor_curvas=self.painel.get_sobreposicao_cfg()
        )

    def _flush_redesenho_pendente(self):
        if self._redraw_after_id is None:
            return
        try:
            self.root.after_cancel(self._redraw_after_id)
        except Exception:
            pass
        self._redesenhar_agora()

    # Grid 

    def _aplicar_cfg_grid(self):
        cfg = self.painel.get_grid_cfg()
        cp = self.canvas_plot
        cp.grid_x         = cfg["grid_x"]
        cp.grid_y         = cfg["grid_y"]
        cp.grid_intervalo = cfg["intervalo"]
        cp.xlim_min       = cfg["xlim_min"]
        cp.xlim_max       = cfg["xlim_max"]

    def on_grid_change(self):
        self._aplicar_cfg_grid()
        self.redesenhar()

    def on_sobreposicao_change(self):
        estado = "ativada" if self.painel.get_sobreposicao_cfg() else "desativada"
        self.painel.status(f"Sobreposicao de curvas {estado}.")
        self.redesenhar()

    # Fontes

    def _aplicar_cfg_fontes(self):
        cfg = self.painel.get_fontes_cfg()
        cp = self.canvas_plot
        cp.fonte_titulo  = cfg["titulo"]
        cp.fonte_eixo    = cfg["eixo"]
        cp.fonte_ticks   = cfg["ticks"]
        cp.fonte_legenda = cfg["legenda"]
        cp.fonte_picos   = cfg["picos"]

    def on_fonte_change(self):
        self._aplicar_cfg_fontes()
        self.redesenhar()

    # Picos 

    def _aplicar_cfg_picos(self):
        cfg = self.painel.get_picos_cfg()
        cp  = self.canvas_plot
        cp.picos_ativos     = cfg["ativo"]
        cp.picos_prominence = cfg["prominence"]
        cp.picos_distancia  = cfg["distancia"]
        cp.picos_max        = cfg["max_picos"]

    def on_picos_toggle(self):
        self._aplicar_cfg_picos()
        self.redesenhar()
        estado = "ativada" if self.canvas_plot.picos_ativos else "desativada"
        self.painel.status(f"Detecção de picos {estado}.")

    def on_picos_cfg_change(self):
        if self.canvas_plot.picos_ativos:
            self._aplicar_cfg_picos()
            self.redesenhar()

    # Tratamentos 

    def abrir_corte(self):
        if not self.amostras:
            messagebox.showinfo("H-DMAPlot", "Carregue arquivos primeiro.")
            return
        DialogSelecionarAmostraColuna(
            self.root, self.amostras, "Corte — selecionar amostra/coluna",
            lambda am, col: DialogCorte(
                self.root, am, col,
                lambda i0, i1: self._aplicar_corte(am, col, i0, i1)
            )
        )

    def _aplicar_corte(self, am: Amostra, col: str, i0: int, i1: int):
        am.cortes[col] = (i0, i1)
        am.invalidar_cache()
        self.redesenhar()
        self.painel.status(f"Corte [{i0}:{i1}] aplicado em '{am.nome}/{col}'.")

    def abrir_offset(self):
        if not self.amostras:
            messagebox.showinfo("H-DMAPlot", "Carregue arquivos primeiro.")
            return
        DialogSelecionarAmostraColuna(
            self.root, self.amostras, "Offset — selecionar amostra/coluna",
            lambda am, col: DialogOffset(
                self.root, am, col,
                lambda val: self._aplicar_offset(am, col, val)
            )
        )

    def _aplicar_offset(self, am: Amostra, col: str, val: float):
        am.offsets[col] = val
        am.invalidar_cache()
        self.redesenhar()
        self.painel.status(f"Offset {val:+.4g} aplicado em '{am.nome}/{col}'.")

    def limpar_marcacoes_fixas(self):
        self.canvas_plot.limpar_marcacoes_fixas()
        self.redesenhar()
        self.painel.status("Marcações fixas removidas.")

    def resetar_tratamentos(self):
        am = self._amostra_ativa()
        if not am:
            return
        am.offsets.clear()
        am.cortes.clear()
        am.suavizacoes.clear()
        am.invalidar_cache()
        self.redesenhar()
        self.painel.status(f"Tratamentos de '{am.nome}' resetados.")

    def aplicar_suavizacao(self):
        if not self.amostras:
            messagebox.showinfo("H-DMAPlot", "Carregue arquivos primeiro.")
            return
        cols_y = self.painel.get_cols_y()
        if not cols_y:
            messagebox.showinfo("H-DMAPlot", "Selecione ao menos um eixo Y antes de suavizar.")
            return
        cfg = self.painel.get_suavizacao_cfg()
        am = self._amostra_ativa()
        if not am:
            return
        for col in cols_y:
            if col in am.df.columns:
                am.suavizacoes[col] = cfg.copy()
        am.invalidar_cache()
        self.redesenhar()
        metodo_nome = "Savitzky-Golay" if cfg["metodo"] == "savgol" else "Média Móvel"
        self.painel.status(
            f"{metodo_nome} (janela={cfg['janela']}) aplicado em '{am.nome}'."
        )

    def remover_suavizacao(self):
        am = self._amostra_ativa()
        if not am:
            return
        am.suavizacoes.clear()
        am.invalidar_cache()
        self.redesenhar()
        self.painel.status(f"Suavização removida de '{am.nome}'.")

    # Exportar 

    def exportar_png(self):
        if not self.amostras:
            messagebox.showinfo("H-DMAPlot", "Nenhum dado disponível.")
            return
        self._flush_redesenho_pendente()
        path = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[("PNG", "*.png")],
            title="Salvar PNG"
        )
        if not path:
            return
        try:
            self.canvas_plot.salvar_png(path, dpi=300)
            self.painel.status(f"PNG salvo: {os.path.basename(path)}")
            messagebox.showinfo("H-DMAPlot", f"PNG exportado com sucesso!\n{path}")
        except Exception as e:
            messagebox.showerror("H-DMAPlot", f"Erro ao salvar PNG:\n{e}")

    def exportar_pdf(self):
        if not self.amostras:
            messagebox.showinfo("H-DMAPlot", "Nenhum dado disponível.")
            return
        self._flush_redesenho_pendente()
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            title="Salvar PDF"
        )
        if not path:
            return
        try:
            # Gera imagem temporária e insere no PDF com ReportLab
            tmp_img = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
            tmp_img.close()
            self.canvas_plot.salvar_png(tmp_img.name, dpi=200)

            doc = SimpleDocTemplate(path)
            styles = getSampleStyleSheet()
            elementos = []
            elementos.append(Paragraph("Relatório H-DMAPlot", styles["Title"]))
            elementos.append(Spacer(1, 10))
            for am in self.amostras:
                elementos.append(Paragraph(f"<b>{am.nome}</b>  — {am.caminho}",
                                           styles["Normal"]))
            elementos.append(Spacer(1, 12))
            elementos.append(RLImage(tmp_img.name, width=460, height=260))
            doc.build(elementos)
            os.unlink(tmp_img.name)
            self.painel.status(f"PDF salvo: {os.path.basename(path)}")
            messagebox.showinfo("H-DMAPlot", f"PDF exportado com sucesso!\n{path}")
        except Exception as e:
            messagebox.showerror("H-DMAPlot", f"Erro ao salvar PDF:\n{e}")

    def exportar_excel(self):
        if not self.amostras:
            messagebox.showinfo("H-DMAPlot", "Nenhum dado disponível.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            title="Salvar Excel"
        )
        if not path:
            return
        try:
            wb = Workbook()
            wb.remove(wb.active)

            header_fill  = PatternFill("solid", fgColor="1E3A5F")
            header_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
            header_align = Alignment(horizontal="center", vertical="center")
            thin_border  = Border(
                left=Side(style="thin", color="C0C0C0"),
                right=Side(style="thin", color="C0C0C0"),
                bottom=Side(style="thin", color="C0C0C0"),
                top=Side(style="thin", color="C0C0C0"),
            )
            alt_fill = PatternFill("solid", fgColor="EEF3FA")
            no_fill = PatternFill()
            body_font = Font(name="Calibri", size=9)

            col_x  = self.painel.get_col_x()
            cols_y = self.painel.get_cols_y()
            export_cols = ([col_x] + [c for c in cols_y if c != col_x]) if col_x else None

            for am in self.amostras:
                ws = wb.create_sheet(title=am.nome[:31])
                df = am.df if export_cols is None else am.df[
                    [c for c in export_cols if c in am.df.columns]
                ]
                df_export = df.round(6)
                for ri, row_data in enumerate(dataframe_to_rows(df_export, index=False, header=True), 1):
                    ws.append(row_data)
                    fill = alt_fill if ri % 2 == 0 else no_fill
                    for cell in ws[ri]:
                        cell.border = thin_border
                        if ri == 1:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_align
                            ws.column_dimensions[cell.column_letter].width = max(14, len(str(cell.value)) + 2)
                        else:
                            cell.font = body_font
                            if fill.patternType:
                                cell.fill = fill

                ws.freeze_panes = "A2"

            # Aba resumo estatístico
            ws_res = wb.create_sheet("Resumo", 0)
            res_headers = ["Amostra", "Coluna", "Mín", "Máx", "Média", "Desvio Padrão", "N pontos"]
            for ci, h in enumerate(res_headers, 1):
                cell = ws_res.cell(1, ci, h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
                ws_res.column_dimensions[cell.column_letter].width = 16

            ri = 2
            for am in self.amostras:
                check_cols = export_cols or am.colunas
                for col in check_cols:
                    if col not in am.df.columns:
                        continue
                    vals = am.df[col].to_numpy(dtype=float)
                    row_vals = [
                        am.nome, col,
                        round(float(np.min(vals)), 4),
                        round(float(np.max(vals)), 4),
                        round(float(np.mean(vals)), 4),
                        round(float(np.std(vals)), 4),
                        len(vals),
                    ]
                    for ci, v in enumerate(row_vals, 1):
                        cell = ws_res.cell(ri, ci, v)
                        cell.border = thin_border
                        cell.font = body_font
                    ri += 1

            ws_res.freeze_panes = "A2"
            wb.save(path)
            self.painel.status(f"Excel salvo: {os.path.basename(path)}")
            messagebox.showinfo("H-DMAPlot", f"Excel exportado com sucesso!\n{path}")
        except Exception as e:
            messagebox.showerror("H-DMAPlot", f"Erro ao salvar Excel:\n{e}")


if __name__ == "__main__":
    App()